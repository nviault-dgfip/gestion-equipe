
import pandas as pd
import numpy as np
from datetime import date
import calendar
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font

# --- CONFIGURATION DE L'ÉQUIPE ---
team_members = [
    "Martin Dupont",
    "Julie Durand",
    "Luc Lefebvre",
    "Sophie Martin",
    "Thomas Petit"
]

year = 2026

# Configuration des textes
header_text = "Liste des Collaborateurs"
footer_text = "Fin de liste des collaborateurs"

mois_fr = {
    1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril", 5: "Mai", 6: "Juin",
    7: "Juillet", 8: "Août", 9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"
}

file_path = 'planning_equipe_format_NN_2026.xlsx'

# --- STYLES ---
grey_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid') # Week-end
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid') # Bleu Entête
text_white_bold = Font(color='FFFFFF', bold=True)
text_italic = Font(italic=True, color='555555')
center_align_vert = Alignment(vertical='center', horizontal='left')

# FORMAT DATE SPÉCIFIQUE DEMANDÉ : "NN J MMM AA"
# Traduction pour Excel/OpenPyXL :
# [$-fr-FR] = Force la langue française
# ddd       = NN (Nom jour court, ex: Jeu.)
# d         = J  (Jour numéro sans zéro, ex: 1)
# mmm       = MMM (Mois court, ex: Janv.)
# yy        = AA (Année 2 chiffres, ex: 26)
excel_date_format = "[$-fr-FR]ddd d mmm yy;@"

with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
   
    # =========================================================================
    # 1. CRÉATION DE L'ONGLET PARAMÈTRES
    # =========================================================================
    col_data = [header_text] + team_members + [footer_text]
    df_equipe = pd.DataFrame({"Data": col_data})
   
    sheet_equipe_name = "Paramètres_Equipe"
    df_equipe.to_excel(writer, sheet_name=sheet_equipe_name, index=False, header=False)
   
    ws_equipe = writer.sheets[sheet_equipe_name]
    ws_equipe.column_dimensions['A'].width = 30
   
    ws_equipe['A1'].fill = header_fill
    ws_equipe['A1'].font = text_white_bold
    ws_equipe['A1'].alignment = Alignment(horizontal='center')
   
    last_row_param = len(col_data)
    ws_equipe.cell(row=last_row_param, column=1).font = text_italic
    ws_equipe.cell(row=last_row_param, column=1).alignment = Alignment(horizontal='center')
   
    # =========================================================================
    # 2. CRÉATION DES ONGLETS MENSUELS
    # =========================================================================
   
    nb_members = len(team_members)
   
    for month in range(1, 13):
       
        num_days = calendar.monthrange(year, month)[1]
        dates_obj = [date(year, month, day) for day in range(1, num_days + 1)]
       
        # Structure de base
        total_cols = 2 + nb_members
        cols_temp = ["Date", "Moment"] + [f"Staff_{i}" for i in range(nb_members)]
       
        # Création DataFrame vide pour la structure
        df = pd.DataFrame(np.full((num_days * 2, total_cols), ''), columns=cols_temp)
       
        sheet_name = f"{mois_fr[month]}_{year}"
        df.to_excel(writer, sheet_name=sheet_name, index=False)
       
        ws = writer.sheets[sheet_name]
       
        # --- A. LIGNE 1 : ENTÊTES ---
        ws['A1'] = "Date"
        ws['B1'] = "Période"
       
        for cell in [ws['A1'], ws['B1']]:
            cell.fill = header_fill
            cell.font = text_white_bold
       
        # Formules pour les noms (C, D, E...)
        for i in range(nb_members):
            col_idx = 3 + i
            col_letter = get_column_letter(col_idx)
            row_source = i + 2
           
            ws.cell(row=1, column=col_idx).value = f"='{sheet_equipe_name}'!A{row_source}"
           
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = text_white_bold
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[col_letter].width = 15

        # --- B. CORPS DU PLANNING ---
        ws.freeze_panes = 'C2'
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
       
        for i, current_date in enumerate(dates_obj):
            row_start = 2 + (i * 2)
            row_end = row_start + 1
           
            # 1. Date (Colonne A) avec le format "NN J MMM AA"
            ws.merge_cells(start_row=row_start, start_column=1, end_row=row_end, end_column=1)
            cell_date = ws.cell(row=row_start, column=1)
            cell_date.value = current_date
           
            # APPLICATION DU FORMAT DATE ICI
            cell_date.number_format = excel_date_format
            cell_date.alignment = center_align_vert
           
            # 2. Période
            ws.cell(row=row_start, column=2).value = "Matin"
            ws.cell(row=row_end, column=2).value = "Après-midi"
           
            # 3. Grisage Week-end
            if current_date.weekday() >= 5:
                for r in [row_start, row_end]:
                    for c in range(1, total_cols + 1):
                        ws.cell(row=r, column=c).fill = grey_fill
           
            # 4. Bordure
            for c in range(1, total_cols + 1):
                ws.cell(row=row_end, column=c).border = Border(bottom=Side(style='thin', color="DDDDDD"))

print(f"Fichier généré avec le format de date 'NN J MMM AA' : {file_path}")

